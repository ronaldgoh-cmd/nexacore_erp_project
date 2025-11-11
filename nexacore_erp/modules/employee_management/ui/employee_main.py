from datetime import datetime, date
import re
import json
import os
import unicodedata

from PySide6.QtCore import Qt, QDate, QObject, QEvent, QDateTime
from PySide6.QtWidgets import (
    QWidget, QTabWidget, QVBoxLayout, QHBoxLayout, QSplitter, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QLineEdit, QComboBox, QFormLayout, QDateEdit, QFileDialog,
    QDialog, QDialogButtonBox, QMessageBox, QSpinBox, QCheckBox, QGroupBox, QGridLayout,
    QScrollArea, QListWidgetItem, QHeaderView, QAbstractItemView, QSizePolicy, QListWidget,
    QInputDialog, QApplication, QDateTimeEdit
)
from PySide6.QtWidgets import QCalendarWidget  # add this near your other QtWidgets imports
from PySide6.QtWidgets import QDoubleSpinBox

# -------- Roles & Access manifest --------
MODULE_KEY = "employee_management"
MODULE_NAME = "Employee Management"
SUBMODULES = [
    ("list",     "Employee List"),
    ("defaults", "Default Leave"),   # change if your tab label differs
    ("settings", "Settings"),        # change if needed
]

def module_manifest() -> dict:
    return {
        "key": MODULE_KEY,
        "name": MODULE_NAME,
        "submodules": [{"key": k, "name": n} for k, n in SUBMODULES],
    }

# ---- shared date helpers ----
MIN_DATE = date(1900, 1, 1)
def _fmt_date(d: date | None) -> str:
    return "" if (not d or d <= MIN_DATE) else d.strftime("%Y-%m-%d")

def _clean_text(s: str) -> str:
    """
    Normalise weird spaces and smart quotes *without* losing characters.
    Helps avoid "??" or unexpected glyphs on import.
    """
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    # replace common smart quotes with ASCII
    s = (s.replace("\u2018", "'").replace("\u2019", "'")
           .replace("\u201c", '"').replace("\u201d", '"')
           .replace("\u00a0", " "))
    # strip zero-width
    s = re.sub(r"[\u200B-\u200D\uFEFF]", "", s)
    return s.strip()


class BlankableDateEdit(QDateEdit):
    """Truly blank until set. Popup calendar defaults to today when blank.
       Clear with Delete, Backspace, Esc, or double-click."""
    class _PopupCalendar(QCalendarWidget):
        def __init__(self, owner):
            super().__init__()
            self._owner = owner

        def showEvent(self, ev):  # focus the calendar on today if the edit is blank
            super().showEvent(ev)
            if getattr(self._owner, "_blank", True):
                self.setSelectedDate(QDate.currentDate())
                self.setFocus()

    def __init__(self, display_fmt: str = "dd/MM/yyyy", *a, **kw):
        self._fmt = display_fmt
        self._blank = True
        super().__init__(*a, **kw)

        self.setCalendarPopup(True)
        self.setCalendarWidget(BlankableDateEdit._PopupCalendar(self))  # custom calendar
        self.setMinimumDate(QDate(1900, 1, 1))
        self.setSpecialValueText(" ")           # render minimum date as blank
        super().setDate(self.minimumDate())     # start blank
        self.setDisplayFormat(self._fmt)
        self.setToolTip("Delete/Backspace/Esc to clear. Double-click to clear.")
        self.dateChanged.connect(self._unblank_on_change)

    def textFromDateTime(self, dt: QDateTime):  # type: ignore[override]
        if getattr(self, "_blank", True):
            return ""
        return QDateTimeEdit.textFromDateTime(self, dt)

    def _unblank_on_change(self, _):
        # stay blank if value equals the special minimum date
        self._blank = (self.date() == self.minimumDate())
        self.update()

    def clear(self):
        self._blank = True
        super().setDate(self.minimumDate())
        self.update()

    def set_real_date(self, qdate: QDate | None):
        if qdate and qdate.isValid():
            self._blank = False
            super().setDate(qdate)
        else:
            self.clear()
        self.update()

    def date_or_none(self):
        return None if self._blank else self.date()

    # QoL: easy ways to blank in Edit mode
    def keyPressEvent(self, ev):  # type: ignore[override]
        if ev.key() in (Qt.Key_Delete, Qt.Key_Backspace, Qt.Key_Escape):
            self.clear()
            ev.accept()
            return
        super().keyPressEvent(ev)

    def mouseDoubleClickEvent(self, ev):  # type: ignore[override]
        self.clear()
        ev.accept()


# optional XLSX support
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None
    DataValidation = None
    get_column_letter = None

from ....core.database import get_employee_session as SessionLocal
from ....core.tenant import id as tenant_id
from ....core.permissions import can_view
from ....core.auth import get_current_user
from ....core.events import employee_events
from ..models import (
    Employee, SalaryHistory, Holiday, DropdownOption, LeaveDefault,
    WorkScheduleDay, LeaveEntitlement
)

# ---- block wheel changes on all combo boxes (when popup closed) ----
class _NoWheelFilter(QObject):
    def eventFilter(self, obj, ev):
        if isinstance(obj, QComboBox):
            if ev.type() == QEvent.Wheel and not obj.view().isVisible():
                return True
            if ev.type() == QEvent.KeyPress and not obj.view().isVisible():
                if ev.key() in (Qt.Key_PageUp, Qt.Key_PageDown):
                    return True
        return False

_NO_WHEEL_FILTER = _NoWheelFilter()
app = QApplication.instance()
if app is not None:
    app.installEventFilter(_NO_WHEEL_FILTER)


# ---------------- Employee code settings (session-scope) ----------------
EMP_CODE_PREFIX = "EM-"
EMP_CODE_ZPAD = 4

# Valid dropdown categories managed in UI
MANAGED_CATEGORIES = [
    "ID Type", "Gender", "Race", "Country", "Residency",
    "Employment Pass", "Department", "Position", "Employment Type", "Bank"
]

DEFAULT_DROPDOWN_OPTIONS: dict[str, list[str]] = {
    "ID Type": ["FIN", "NRIC"],
    "Gender": ["Male", "Female"],
    "Race": ["Chinese", "Eurasian", "Indian", "Malay", "Others"],
    "Employment Type": ["Casual", "Part Time", "Full Time"],
    "Residency": ["Citizen", "Permanent Resident"],
}


def _ensure_dropdown_defaults(category: str) -> None:
    defaults = DEFAULT_DROPDOWN_OPTIONS.get(category)
    if not defaults:
        return
    with SessionLocal() as s:
        existing = {
            row[0]
            for row in s.query(DropdownOption.value)
            .filter(
                DropdownOption.account_id == tenant_id(),
                DropdownOption.category == category,
            )
            .all()
        }
        missing = [v for v in defaults if v not in existing]
        if not missing:
            return
        for value in missing:
            s.add(
                DropdownOption(
                    account_id=tenant_id(),
                    category=category,
                    value=value,
                )
            )
        s.commit()
# persist employee code format to a small json file next to this module
SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "employee_settings.json")

def _load_code_settings():
    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            d = json.load(f)
            p = str(d.get("prefix", "EM-") or "EM-")
            z = int(d.get("zpad", 4) or 4)
            return p, z
    except Exception:
        return "EM-", 4

def _save_code_settings(prefix: str, zpad: int):
    try:
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump({"prefix": prefix, "zpad": int(zpad)}, f)
    except Exception:
        pass

# Columns for Employee List table (order matters)
COLS = [
    ("employment_status", "Status"),
    ("code", "Employee Code"),
    ("full_name", "Employee Name"),
    ("department", "Department"),
    ("position", "Position"),
    ("employment_type", "Employment Type"),
    ("dob", "Date of Birth"),
    ("age", "Age"),
    ("id_type", "Identification Type"),
    ("id_number", "Identification Number"),
    ("country", "Country"),
    ("residency", "Residency"),
    ("join_date", "Join Date"),
    ("exit_date", "Exit Date"),
]


class EmployeeMainWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName("EmployeeMainWidget")
        global EMP_CODE_PREFIX, EMP_CODE_ZPAD
        EMP_CODE_PREFIX, EMP_CODE_ZPAD = _load_code_settings()

        # (defensive) install global no-wheel filter
        app = QApplication.instance()
        if app is not None:
            app.installEventFilter(_NO_WHEEL_FILTER)

        for category in DEFAULT_DROPDOWN_OPTIONS:
            _ensure_dropdown_defaults(category)

        self.tabs = QTabWidget(self)
        v = QVBoxLayout(self)
        v.addWidget(self.tabs)

        added = False
        if self._allowed("Employee List"):
            self._build_employee_list_tab(); added = True
        if self._allowed("Holidays"):
            self._build_holidays_tab(); added = True
        if self._allowed("Employee Settings"):
            self._build_settings_tab(); added = True

        if not added:
            self._no_access_placeholder()

    MODULE_KEY = "employee_management"

    def filter_tabs_by_access(self, allowed_keys: list[str] | set[str]):
        allowed = set(allowed_keys or [])
        if not allowed:
            return
        label_by_key = {
            "list": "Employee List",
            "defaults": "Default Leave",
            "settings": "Settings",
        }
        allowed_labels = {label_by_key[k] for k in allowed if k in label_by_key}
        for i in range(self.tabs.count() - 1, -1, -1):
            if self.tabs.tabText(i) not in allowed_labels:
                self.tabs.removeTab(i)

    def _allowed(self, sub_label: str) -> bool:
        try:
            u = get_current_user()
            if not u:
                return True  # avoid dead UI if unauthenticated context
            if getattr(u, "role", "") == "superadmin" or getattr(u, "id", -1) < 0:
                return True
            return can_view(u.id, "Employee Management", None, sub_label)
        except Exception:
            return True  # fail-open to avoid breaking UI

    def _no_access_placeholder(self):
        w = QWidget()
        v = QVBoxLayout(w)
        msg = QLabel("No access to any Employee Management submodule.")
        msg.setAlignment(Qt.AlignCenter)
        v.addWidget(msg)
        self.tabs.addTab(w, "Access")

    # optional: lets module.py switch to a target tab after constructing this widget
    def select_tab(self, label: str):
        target = (label or "").strip().lower()
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i).strip().lower() == target:
                self.tabs.setCurrentIndex(i)
                return

    # small helper: fetch dropdown option list for a category
    def _opts(self, category: str) -> list[str]:
        _ensure_dropdown_defaults(category)
        with SessionLocal() as s:
            rows = s.query(DropdownOption)\
                    .filter(DropdownOption.account_id == tenant_id(),
                            DropdownOption.category == category)\
                    .order_by(DropdownOption.value).all()
        return [r.value for r in rows]

    # ---------------- Employee List ----------------
    def _build_employee_list_tab(self):
        host = QWidget()
        h = QHBoxLayout(host)
        splitter = QSplitter(Qt.Horizontal, host)

        # left: table + filters + actions
        left = QWidget(splitter)
        lv = QVBoxLayout(left)

        # actions row
        actions = QHBoxLayout()
        self.filter_toggle = QPushButton("Filters ▸")
        self.filter_toggle.setCheckable(True)
        self.filter_toggle.setChecked(False)
        self.filter_toggle.toggled.connect(self._toggle_filters)
        actions.addWidget(self.filter_toggle)

        self.quick_search = QLineEdit()
        self.quick_search.setPlaceholderText("Quick search (any column)…")
        self.quick_search.textChanged.connect(self._apply_filters)
        btn_add = QPushButton("Add"); btn_add.clicked.connect(self._add_employee)
        btn_edit = QPushButton("Edit"); btn_edit.clicked.connect(self._edit_employee)
        btn_del = QPushButton("Delete"); btn_del.clicked.connect(self._delete_employee)
        for w in (self.quick_search, btn_add, btn_edit, btn_del):
            actions.addWidget(w)
        lv.addLayout(actions)

        # per-column filters
        self.filter_box = QGroupBox("Filters")
        fbv = QVBoxLayout(self.filter_box)
        grid = QGridLayout()
        self.filters: dict[str, QWidget] = {}

        def add_filter(row, label, key, widget_factory):
            grid.addWidget(QLabel(label), row, 0)
            w = widget_factory()
            self.filters[key] = w
            if isinstance(w, QLineEdit):
                w.textChanged.connect(self._apply_filters)
            elif isinstance(w, QComboBox):
                w.currentTextChanged.connect(self._apply_filters)
            elif isinstance(w, QSpinBox):
                w.valueChanged.connect(self._apply_filters)
            grid.addWidget(w, row, 1)

        # dropdown helpers with blank
        def mk_dd(values: list[str]) -> QComboBox:
            cb = QComboBox()
            cb.addItem("")
            for v in values or []:
                cb.addItem(v)
            return cb

        r = 0
        add_filter(r, "Status", "employment_status", lambda: mk_dd(["Active", "Non-Active"])); r += 1
        add_filter(r, "Employee Code", "code", QLineEdit); r += 1
        add_filter(r, "Employee Name", "full_name", QLineEdit); r += 1
        add_filter(r, "Department", "department", lambda: mk_dd(self._opts("Department"))); r += 1
        add_filter(r, "Position", "position",   lambda: mk_dd(self._opts("Position"))); r += 1
        add_filter(r, "Employment Type", "employment_type", lambda: mk_dd(self._opts("Employment Type") or ["Full-Time", "Part-Time", "Contract"])); r += 1
        add_filter(r, "ID Type", "id_type",     lambda: mk_dd(self._opts("ID Type"))); r += 1
        add_filter(r, "ID Number", "id_number", QLineEdit); r += 1
        add_filter(r, "Country", "country",     lambda: mk_dd(self._opts("Country"))); r += 1
        add_filter(r, "Residency", "residency", lambda: mk_dd(self._opts("Residency") or ["Citizen", "Permanent Resident", "Work Pass"])); r += 1

        # Age numeric filters (not dropdown)
        def mk_age():
            sp = QSpinBox()
            sp.setRange(0, 120)
            sp.setSpecialValueText("")  # 0 acts as blank
            sp.setValue(0)
            return sp
        add_filter(r, "Age ≥", "age_min", mk_age); r += 1
        add_filter(r, "Age ≤", "age_max", mk_age); r += 1

        fbv.addLayout(grid)

        self.filter_area = QScrollArea()
        self.filter_area.setWidget(self.filter_box)
        self.filter_area.setWidgetResizable(True)
        self.filter_area.setFixedHeight(180)
        self.filter_area.setVisible(False)
        lv.addWidget(self.filter_area)

        # table
        self.emp_table = QTableWidget(0, len(COLS))
        self.emp_table.setHorizontalHeaderLabels([c[1] for c in COLS])
        self.emp_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.emp_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.emp_table.setSortingEnabled(False)
        self.emp_table.itemSelectionChanged.connect(self._show_preview)
        hdr = self.emp_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setStretchLastSection(False)
        hdr.setSortIndicatorShown(False)
        hdr.setSectionsClickable(False)
        lv.addWidget(self.emp_table, 1)

        # right: detail
        right = QWidget(splitter)
        rv = QVBoxLayout(right)
        self.detail_form = self._build_readonly_form()
        rv.addWidget(self.detail_form["host"], 1)

        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)

        h.addWidget(splitter)
        self.tabs.addTab(host, "Employee List")
        self._reload_employees()

    def _toggle_filters(self, checked: bool):
        self.filter_area.setVisible(checked)
        self.filter_toggle.setText("Filters ▾" if checked else "Filters ▸")

    def _combo_with(self, items: list[str]) -> QComboBox:
        c = QComboBox()
        c.addItems(items)
        return c

    def _reload_employees(self):
        with SessionLocal() as s:
            rows = s.query(Employee).filter(Employee.account_id == tenant_id()).all()
        self._all_rows = rows
        self._apply_filters()

    def _notify_employees_changed(self):
        try:
            employee_events.employees_changed.emit()
        except Exception:
            pass

    def _apply_filters(self):
        txt = (self.quick_search.text() or "").lower().strip()

        def f_get(k):
            w = self.filters.get(k)
            if isinstance(w, QComboBox):
                return w.currentText().strip().lower()
            if isinstance(w, QLineEdit):
                return w.text().strip().lower()
            if isinstance(w, QSpinBox):
                return "" if w.value() == 0 else str(w.value())
            return ""

        def to_int(val):
            try:
                return int(val) if val != "" else None
            except Exception:
                return None

        age_min = to_int(f_get("age_min"))
        age_max = to_int(f_get("age_max"))

        def keep(e: Employee) -> bool:
            if not ((e.code or "").strip() or (e.full_name or "").strip()):
                return False
            if txt:
                hay = " ".join([
                    str(e.employment_status or ""), str(e.code or ""), str(e.full_name or ""),
                    str(getattr(e, "department", "") or ""), str(e.position or ""),
                    str(e.employment_type or ""), str(e.id_type or ""), str(e.id_number or ""),
                    str(e.country or ""), str(e.residency or "")
                ]).lower()
                if txt not in hay:
                    return False
            for key in ["employment_status", "code", "full_name", "department", "position",
                        "employment_type", "id_type", "id_number", "country", "residency"]:
                val = f_get(key)
                if val and val not in str(getattr(e, key, "") or "").lower():
                    return False
            age = None
            if e.dob and e.dob > MIN_DATE:
                try:
                    age = int((datetime.utcnow().date() - e.dob).days // 365.25)
                except Exception:
                    age = None
            if age_min is not None and (age is None or age < age_min):
                return False
            if age_max is not None and (age is None or age > age_max):
                return False
            return True

        rows = [e for e in getattr(self, "_all_rows", []) if keep(e)]

        self.emp_table.setSortingEnabled(False)
        self.emp_table.setRowCount(len(rows))
        for r, e in enumerate(rows):
            def put(col, val):
                self.emp_table.setItem(r, col, QTableWidgetItem(val))

            employment_status = e.employment_status or ""
            code = e.code or ""
            full_name = e.full_name or ""
            department = getattr(e, "department", "") or ""
            position = e.position or ""
            employment_type = e.employment_type or ""
            dob_txt = _fmt_date(e.dob)
            age_txt = ""
            if e.dob and e.dob > MIN_DATE:
                try:
                    age_txt = str(int((datetime.utcnow().date() - e.dob).days // 365.25))
                except Exception:
                    age_txt = ""
            id_type = e.id_type or ""
            id_number = e.id_number or ""
            country = e.country or ""
            residency = e.residency or ""
            join_date = _fmt_date(e.join_date)
            exit_date = _fmt_date(e.exit_date)

            values = [employment_status, code, full_name, department, position, employment_type,
                      dob_txt, age_txt, id_type, id_number, country, residency, join_date, exit_date]
            for c, v in enumerate(values):
                put(c, v)

            self.emp_table.setVerticalHeaderItem(r, QTableWidgetItem(str(e.id)))

        if self.emp_table.rowCount() > 0:
            self.emp_table.selectRow(0)
        else:
            self._clear_preview()

    # read-only detail form
    def _build_readonly_form(self):
        host = QWidget()
        v = QVBoxLayout(host)
        frm = QFormLayout()
        fields = [
            "Employee Code", "Full Name", "Email", "Contact Number", "Address",
            "ID Type", "ID Number", "Gender", "Date of Birth", "Age", "Race",
            "Country", "Residency", "PR Date",
            "Employment Status", "Employment Pass", "Work Permit Number",
            "Department", "Position", "Employment Type",
            "Join Date", "Exit Date", "Holiday Group",
            "Bank", "Account Number",
            "Basic Salary", "Incentives", "Allowance", "Overtime Rate", "Part Time Rate", "Levy"
        ]
        labels = {}
        for f in fields:
            lbl = QLabel("")
            lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
            frm.addRow(f, lbl)
            labels[f] = lbl
        v.addLayout(frm)
        v.addStretch(1)
        return {"host": host, "labels": labels}

    def _clear_preview(self):
        for lbl in self.detail_form["labels"].values():
            lbl.setText("")

    def _show_preview(self):
        sel = self.emp_table.selectedItems()
        if not sel:
            self._clear_preview()
            return
        r = self.emp_table.currentRow()
        id_item = self.emp_table.verticalHeaderItem(r)
        emp_id = int(id_item.text()) if id_item else None
        if not emp_id:
            self._clear_preview()
            return
        with SessionLocal() as s:
            e = s.get(Employee, emp_id)
        if not e:
            self._clear_preview()
            return

        L = self.detail_form["labels"]
        fmt = _fmt_date
        age_txt = ""
        if e.dob and e.dob > MIN_DATE:
            try:
                age_txt = str(int((datetime.utcnow().date() - e.dob).days // 365.25))
            except Exception:
                age_txt = ""

        def put(k, v):
            if isinstance(v, str):
                L[k].setText(v)
            elif isinstance(v, float):
                L[k].setText(f"{v:.2f}")
            else:
                L[k].setText("" if v is None else str(v))

        put("Employee Code", e.code or "")
        put("Full Name", e.full_name or "")
        put("Email", e.email or "")
        put("Contact Number", e.contact_number or "")
        put("Address", e.address or "")
        put("ID Type", e.id_type or "")
        put("ID Number", e.id_number or "")
        put("Gender", e.gender or "")
        put("Date of Birth", fmt(e.dob))
        put("Age", age_txt)
        put("Race", e.race or "")
        put("Country", e.country or "")
        put("Residency", e.residency or "")
        put("PR Date", fmt(e.pr_date))
        put("Employment Status", e.employment_status or "")
        put("Employment Pass", e.employment_pass or "")
        put("Work Permit Number", e.work_permit_number or "")
        put("Department", getattr(e, "department", "") or "")
        put("Position", e.position or "")
        put("Employment Type", e.employment_type or "")
        put("Join Date", fmt(e.join_date))
        put("Exit Date", fmt(e.exit_date))
        put("Holiday Group", e.holiday_group or "")
        put("Bank", e.bank or "")
        put("Account Number", e.bank_account or "")
        put("Basic Salary", e.basic_salary or 0.0)
        put("Incentives", e.incentives or 0.0)
        put("Allowance", e.allowance or 0.0)
        put("Overtime Rate", e.overtime_rate or 0.0)
        put("Part Time Rate", e.parttime_rate or 0.0)
        put("Levy", e.levy or 0.0)

    # --- list actions ---
    def _add_employee(self):
        d = EmployeeEditor(parent=self)
        if d.exec() == QDialog.Accepted:
            self._reload_employees()
            self._notify_employees_changed()

    def _edit_employee(self):
        r = self.emp_table.currentRow()
        if r < 0:
            return
        id_item = self.emp_table.verticalHeaderItem(r)
        emp_id = int(id_item.text()) if id_item else None
        if not emp_id:
            return
        d = EmployeeEditor(emp_id, parent=self)
        if d.exec() == QDialog.Accepted:
            self._reload_employees()
            for i in range(self.emp_table.rowCount()):
                ii = self.emp_table.verticalHeaderItem(i)
                if ii and ii.text() == str(emp_id):
                    self.emp_table.selectRow(i)
                    break
            self._notify_employees_changed()

    def _delete_employee(self):
        r = self.emp_table.currentRow()
        if r < 0:
            return
        id_item = self.emp_table.verticalHeaderItem(r)
        emp_id = int(id_item.text()) if id_item else None
        if not emp_id:
            return
        if QMessageBox.question(self, "Delete", "Delete selected employee?") != QMessageBox.Yes:
            return
        with SessionLocal() as s:
            e = s.get(Employee, emp_id)
            if e:
                s.delete(e)
                s.commit()
        self._reload_employees()
        self._notify_employees_changed()

    # ---------------- Holidays ----------------
    def _build_holidays_tab(self):
        host = QWidget(); v = QVBoxLayout(host)
        top = QHBoxLayout()
        self.h_group = QLineEdit(); self.h_group.setPlaceholderText("Holiday Group e.g. A or B")
        self.h_name = QLineEdit(); self.h_name.setPlaceholderText("Holiday Name")
        self.h_date = QLineEdit(); self.h_date.setPlaceholderText("Date DD-MM-YYYY")
        btn_add = QPushButton("Add"); btn_add.clicked.connect(self._holiday_add)
        btn_del = QPushButton("Delete Selected"); btn_del.clicked.connect(self._holiday_del)
        btn_imp = QPushButton("Import CSV"); btn_imp.clicked.connect(self._holiday_import_csv)
        btn_tpl = QPushButton("Download Template"); btn_tpl.clicked.connect(self._holiday_template)
        for w in (self.h_group, self.h_name, self.h_date, btn_add, btn_del, btn_imp, btn_tpl):
            top.addWidget(w)
        v.addLayout(top)

        self.h_table = QTableWidget(0, 3)
        self.h_table.setHorizontalHeaderLabels(["Group", "Name", "Date"])
        hdr = self.h_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setStretchLastSection(False)
        self.h_table.verticalHeader().setVisible(False)

        v.addWidget(self.h_table, 1)

        self.tabs.addTab(host, "Holidays")
        self._holiday_reload()

    def _holiday_reload(self):
        with SessionLocal() as s:
            rows = (
                s.query(Holiday)
                .filter(Holiday.account_id == tenant_id())
                .order_by(Holiday.group_code, Holiday.date)
                .all()
            )

        self.h_table.setRowCount(0)
        for r in rows:
            row = self.h_table.rowCount()
            self.h_table.insertRow(row)
            self.h_table.setItem(row, 0, QTableWidgetItem(r.group_code))
            self.h_table.setItem(row, 1, QTableWidgetItem(r.name))
            self.h_table.setItem(row, 2, QTableWidgetItem(r.date.strftime("%d-%m-%Y")))

        self.h_table.resizeColumnsToContents()

    def _holiday_add(self):
        g = _clean_text(self.h_group.text())
        n = _clean_text(self.h_name.text())
        d = _clean_text(self.h_date.text())
        if not (g and n and d):
            QMessageBox.warning(self, "Holiday", "Fill group, name, date (DD-MM-YYYY)"); return
        try:
            day, month, year = map(int, d.split("-")); dt = date(year, month, day)
        except Exception:
            QMessageBox.warning(self, "Holiday", "Date format must be DD-MM-YYYY"); return
        with SessionLocal() as s:
            s.add(Holiday(account_id=tenant_id(), group_code=g, name=n, date=dt))
            try: s.commit()
            except Exception as ex:
                s.rollback(); QMessageBox.warning(self, "Holiday", f"Duplicate or invalid: {ex}")
        self._holiday_reload()

    def _holiday_del(self):
        rows = sorted({r.row() for r in self.h_table.selectedIndexes()}, reverse=True)
        if not rows: return
        items = []
        for r in rows:
            items.append((self.h_table.item(r, 0).text(), self.h_table.item(r, 1).text(), self.h_table.item(r, 2).text()))
        with SessionLocal() as s:
            for g, n, d in items:
                day, month, year = map(int, d.split("-")); dt = date(year, month, day)
                q = s.query(Holiday).filter(Holiday.account_id == tenant_id(), Holiday.group_code == g, Holiday.name == n, Holiday.date == dt)
                for h in q.all(): s.delete(h)
            s.commit()
        self._holiday_reload()

    def _holiday_import_csv(self):
        import csv, io

        path, _ = QFileDialog.getOpenFileName(self, "Import Holidays CSV", "", "CSV Files (*.csv)")
        if not path:
            return

        raw = open(path, "rb").read()

        def _try_decode(b: bytes, enc: str):
            try:
                return b.decode(enc)
            except UnicodeDecodeError:
                return None

        text = None
        used_enc = None
        for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            s = _try_decode(raw, enc)
            if s is not None:
                text, used_enc = s, enc
                break
        if text is None:
            text, used_enc = raw.decode("latin-1", "replace"), "latin-1(replace)"

        # normalise once here
        text = unicodedata.normalize("NFKC", text)
        text = (text.replace("\u00a0", " ")
                    .replace("\u2018", "'").replace("\u2019", "'")
                    .replace("\u201c", '"').replace("\u201d", '"'))
        text = re.sub(r"[\u200B-\u200D\uFEFF]", "", text)

        # sniff delimiter
        try:
            sample = text[:4096]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delim = dialect.delimiter
        except Exception:
            delim = ","

        # header normaliser
        def norm_key(k: str) -> str:
            k = unicodedata.normalize("NFKC", k or "").strip().lower()
            k = re.sub(r"\s+", " ", k)
            k = k.replace("group code", "group")
            k = re.sub(r"[^a-z0-9]+", "", k)
            return k

        key_map = {
            "group": "group",
            "groupcode": "group",
            "grp": "group",
            "name": "name",
            "holiday": "name",
            "holidayname": "name",
            "date": "date",
            "holidaydate": "date",
            "dt": "date",
            "ishalfday": "is_half_day",
            "description": "description",
            "desc": "description",
        }

        rd = csv.reader(io.StringIO(text), delimiter=delim)
        try:
            raw_headers = next(rd)
        except StopIteration:
            QMessageBox.warning(self, "Holidays", "Empty CSV.")
            return
        headers = [key_map.get(norm_key(h), norm_key(h)) for h in raw_headers]

        def idx_of(field):
            try:
                return headers.index(field)
            except ValueError:
                return -1

        i_g = idx_of("group")
        i_n = idx_of("name")
        i_d = idx_of("date")

        if min(i_g, i_n, i_d) < 0:
            QMessageBox.warning(
                self, "Holidays",
                "Missing required headers. Need columns that map to: group, name, date.\n"
                f"Detected headers: {headers}\nEncoding: {used_enc}, delimiter: '{delim}'"
            )
            return

        def parse_date(s: str):
            from datetime import timedelta
            s = _clean_text(s)
            if not s:
                return None

            # Excel serial number
            if re.fullmatch(r"\d{1,6}", s):
                try:
                    return (datetime(1899, 12, 30) + timedelta(days=int(s))).date()
                except Exception:
                    pass

            s = s.replace(",", "").replace(".", "")
            s = re.sub(r"\s+", " ", s)

            fmts = (
                "%Y-%m-%d",
                "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y",
                "%d-%m-%y", "%d/%m/%y", "%m/%d/%y",
                "%d-%b-%Y", "%d %b %Y", "%d-%B-%Y", "%d %B %Y",
                "%d-%b-%y", "%d %b %y", "%d-%B-%y", "%d %B %y",
            )
            for fmt in fmts:
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    continue
            return None

        inserted, skipped = 0, 0
        reasons = []

        with SessionLocal() as s:
            for row in rd:
                if len(row) < len(headers):
                    row += [""] * (len(headers) - len(row))

                g = _clean_text(row[i_g] if i_g >= 0 else "")
                n = _clean_text(row[i_n] if i_n >= 0 else "")
                d = _clean_text(row[i_d] if i_d >= 0 else "")

                if not g or not n or not d:
                    skipped += 1
                    if len(reasons) < 5:
                        reasons.append(f"Missing field(s): group='{g}', name='{n}', date='{d}'")
                    continue

                dt = parse_date(d)
                if not dt:
                    skipped += 1
                    if len(reasons) < 5:
                        reasons.append(f"Unparsed date '{d}'")
                    continue

                try:
                    s.add(Holiday(account_id=tenant_id(), group_code=g, name=n, date=dt))
                    s.flush()
                    inserted += 1
                except Exception as ex:
                    s.rollback()
                    skipped += 1
                    if len(reasons) < 5:
                        reasons.append(f"DB reject for '{g}-{n}-{dt}': {ex}")

            s.commit()

        detail = ("\nReasons (first 5):\n- " + "\n- ".join(reasons)) if reasons else ""
        QMessageBox.information(
            self, "Holidays",
            f"Encoding: {used_enc}\nDelimiter: '{delim}'\nImported: {inserted}\nSkipped: {skipped}{detail}"
        )
        self._holiday_reload()

    def _holiday_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save CSV Template", "holiday_template.csv", "CSV Files (*.csv)")
        if not path: return
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            wr = csv.writer(f)
            wr.writerow(["group", "name", "date"])
            wr.writerow(["A", "New Year's Day", "01-01-2025"])
        QMessageBox.information(self, "Template", "Template saved.")

    # ---------------- Settings ----------------
    def _build_settings_tab(self):
        host = QWidget()
        host.setMaximumWidth(520)
        v = QVBoxLayout(host)

        # Employee ID format
        frm = QFormLayout()
        self.id_prefix = QLineEdit()
        self.zero_pad = QSpinBox()
        self.zero_pad.setRange(2, 8)
        self.zero_pad.setValue(EMP_CODE_ZPAD)
        self.id_prefix.setText(EMP_CODE_PREFIX)

        frm.addRow("Employee ID Prefix", self.id_prefix)
        frm.addRow("Zero Padding", self.zero_pad)
        self.id_preview = QLabel(f"Preview: {EMP_CODE_PREFIX}{1:0{EMP_CODE_ZPAD}d}")
        btn_apply = QPushButton("Apply")
        btn_apply.clicked.connect(self._apply_id_format)
        v.addLayout(frm)
        v.addWidget(self.id_preview)
        v.addWidget(btn_apply)

        v.addSpacing(8)

        # Dropdowns and leave defaults
        row1 = QHBoxLayout()
        btn_opts = QPushButton("Manage Dropdown Options")
        btn_opts.clicked.connect(lambda: DropdownOptionsDialog(self).exec())
        btn_leave = QPushButton("Manage Leave Defaults")
        btn_leave.clicked.connect(lambda: LeaveDefaultsDialog(self).exec())
        row1.addWidget(btn_opts)
        row1.addWidget(btn_leave)
        row1.addStretch(1)
        v.addLayout(row1)

        v.addSpacing(8)

        # Data Management
        row2 = QHBoxLayout()
        exp_btn = QPushButton("Export XLSX")
        exp_btn.clicked.connect(self._export_xlsx)
        imp_btn = QPushButton("Import XLSX")
        imp_btn.clicked.connect(self._import_xlsx)
        tmpl_btn = QPushButton("Download Template")
        tmpl_btn.clicked.connect(self._export_employees_template)
        row2.addWidget(exp_btn)
        row2.addWidget(imp_btn)
        row2.addWidget(tmpl_btn)
        row2.addStretch(1)
        v.addLayout(row2)

        v.addStretch(1)
        self.tabs.addTab(host, "Employee Settings")

    def _export_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
        except Exception:
            QMessageBox.warning(self, "Export", "openpyxl not installed")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Employees", "employees.xlsx", "Excel (*.xlsx)")
        if not path: return

        with SessionLocal() as s:
            drop = {}
            for r in s.query(DropdownOption).filter(DropdownOption.account_id == tenant_id()).all():
                drop.setdefault(r.category, []).append(r.value)
            emps = s.query(Employee).filter(Employee.account_id == tenant_id()).all()
            groups = [g[0] for g in s.query(Holiday.group_code).filter(Holiday.account_id == tenant_id()).distinct().all()]
        drop.setdefault("Holiday Group", groups)
        drop.setdefault("Employment Status", ["Active", "Non-Active"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        ws2 = wb.create_sheet("Dropdowns")

        # write dropdowns
        drow = 1
        for cat, vals in sorted(drop.items()):
            ws2.cell(drow, 1, cat)
            for i, v in enumerate(sorted(vals), start=2):
                ws2.cell(drow, i, _clean_text(v))
            drow += 1

        headers = [
            "Employee Code", "Full Name", "Email", "Contact Number", "Address",
            "ID Type", "ID Number", "Gender", "Date of Birth", "Race", "Country", "Residency", "PR Date",
            "Employment Status", "Employment Pass", "Work Permit Number",
            "Department", "Position", "Employment Type",
            "Join Date", "Exit Date", "Holiday Group",
            "Bank", "Account Number",
            "Incentives", "Allowance", "Overtime Rate", "Part Time Rate", "Levy", "Basic Salary"
        ]
        ws.append(headers)

        for e in emps:
            ws.append([
                e.code or "", e.full_name or "", e.email or "", e.contact_number or "", e.address or "",
                e.id_type or "", e.id_number or "", e.gender or "", _fmt_date(e.dob), e.race or "", e.country or "",
                e.residency or "", _fmt_date(e.pr_date),
                e.employment_status or "", e.employment_pass or "", e.work_permit_number or "",
                getattr(e, "department", "") or "", e.position or "", e.employment_type or "",
                _fmt_date(e.join_date), _fmt_date(e.exit_date), e.holiday_group or "",
                e.bank or "", e.bank_account or "",
                e.incentives or 0.0, e.allowance or 0.0, e.overtime_rate or 0.0, e.parttime_rate or 0.0, e.levy or 0.0,
                e.basic_salary or 0.0
            ])

        col_index = {h: i + 1 for i, h in enumerate(headers)}
        dv_map = {
            "ID Type": "ID Type", "Gender": "Gender", "Race": "Race", "Country": "Country",
            "Residency": "Residency", "Employment Status": "Employment Status",
            "Employment Pass": "Employment Pass", "Department": "Department",
            "Position": "Position", "Employment Type": "Employment Type",
            "Bank": "Bank", "Holiday Group": "Holiday Group",
        }
        cat_formula = {}
        for r in range(1, ws2.max_row + 1):
            cat = ws2.cell(r, 1).value
            if not cat: continue
            cat_formula[cat] = f'=OFFSET(Dropdowns!$B${r},0,0,1,COUNTA(Dropdowns!$B${r}:$ZZ${r}))'

        max_rows = max(1000, ws.max_row + 100)
        for header, cat in dv_map.items():
            if cat not in cat_formula or header not in col_index: continue
            c = col_index[header]
            dv = DataValidation(type="list", formula1=cat_formula[cat], allow_blank=True, showDropDown=True)
            ws.add_data_validation(dv)
            dv.ranges.add(f"{ws.cell(2, c).coordinate}:{ws.cell(max_rows, c).coordinate}")

        ws.freeze_panes = "A2"
        try:
            wb.save(path)
            QMessageBox.information(self, "Export", "Export complete.")
        except Exception as ex:
            QMessageBox.warning(self, "Export", f"Failed to save: {ex}")

    def _import_xlsx(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            QMessageBox.warning(self, "Import", "openpyxl not installed")
            return

        path, _ = QFileDialog.getOpenFileName(self, "Import Employees", "", "Excel (*.xlsx)")
        if not path:
            return

        try:
            wb = load_workbook(path, data_only=True)
            ws = wb["Employees"]
        except Exception as ex:
            QMessageBox.warning(self, "Import", f"Cannot open workbook: {ex}")
            return

        def _norm_header(s: str) -> str:
            if not isinstance(s, str):
                return ""
            s = unicodedata.normalize("NFKC", s)
            s = re.sub(r"[\u00A0\u200B-\u200D\uFEFF]", "", s)
            return s.strip().lower()

        headers: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            key = _norm_header(h) if isinstance(h, str) else ""
            if key:
                headers[key] = c

        if "full name" not in headers and "employee code" not in headers:
            QMessageBox.warning(self, "Import", "Sheet needs 'Full Name' or 'Employee Code'.")
            return

        def gv(row, *names):
            for name in names:
                c = headers.get(_norm_header(name))
                if c:
                    return ws.cell(row, c).value
            return None

        def to_date(x):
            BLANKS = {None, "", "-", "--", "—", "N/A", "NA"}
            if x in BLANKS:
                return None
            if isinstance(x, datetime):
                d = x.date()
                return None if d <= MIN_DATE else d
            if isinstance(x, date):
                return None if x <= MIN_DATE else x
            if isinstance(x, (int, float)):
                try:
                    from openpyxl.utils.datetime import from_excel, CALENDAR_WINDOWS_1900 as CAL
                    d = from_excel(x, CAL).date()
                    return None if d <= MIN_DATE else d
                except Exception:
                    pass
            if isinstance(x, str):
                s = _clean_text(x)
                if re.fullmatch(r"\d{1,6}", s):
                    try:
                        from openpyxl.utils.datetime import from_excel, CALENDAR_WINDOWS_1900 as CAL
                        d = from_excel(int(s), CAL).date()
                        return None if d <= MIN_DATE else d
                    except Exception:
                        pass
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                    try:
                        d = datetime.strptime(s, fmt).date()
                        return None if d <= MIN_DATE else d
                    except Exception:
                        continue
                return None
            return None

        def to_float(x):
            try:
                return float(x)
            except Exception:
                return 0.0

        def to_str(x):
            s = "" if x is None else str(x)
            s = _clean_text(s)
            return "" if s in ("-", "--") else s

        created = updated = 0
        with SessionLocal() as s:
            prefix, z = EMP_CODE_PREFIX, EMP_CODE_ZPAD
            existing_codes = [
                c for (c,) in s.query(Employee.code)
                .filter(Employee.account_id == tenant_id(),
                        Employee.code.isnot(None),
                        Employee.code.like(f"{prefix}%"))
                .all()
            ]

            def _num_tail(c):
                if not c: return None
                m = re.search(r"(\d+)$", c)
                return int(m.group(1)) if m else None

            used_nums = {n for n in (_num_tail(c) for c in existing_codes) if n is not None}
            cursor = max(used_nums) if used_nums else 0

            def next_code():
                nonlocal cursor
                cursor += 1
                code = f"{prefix}{cursor:0{z}d}"
                used_nums.add(cursor)
                return code

            for r in range(2, ws.max_row + 1):
                if not any((ws.cell(r, c).value not in (None, "", " ")) for c in range(1, ws.max_column + 1)):
                    continue

                code_raw = gv(r, "Employee Code", "code")
                name_raw = gv(r, "Full Name", "full_name")

                code = to_str(code_raw)
                full_name = to_str(name_raw)

                if not code and not full_name:
                    continue

                q = s.query(Employee).filter(Employee.account_id == tenant_id())
                e = None
                if code:
                    e = q.filter(Employee.code == code).first()
                if e is None and full_name:
                    e = q.filter(Employee.full_name == full_name).first()

                is_new = e is None
                if is_new:
                    if not code:
                        code = next_code()
                    e = Employee(account_id=tenant_id(), code=code, full_name=full_name)
                    s.add(e)
                else:
                    if code and not (e.code or "").strip():
                        e.code = code
                    if full_name:
                        e.full_name = full_name

                e.email = to_str(gv(r, "Email", "email"))
                e.contact_number = to_str(gv(r, "Contact Number", "contact_number"))
                e.address = to_str(gv(r, "Address", "address"))
                e.id_type = to_str(gv(r, "ID Type", "id_type"))
                e.id_number = to_str(gv(r, "ID Number", "id_number"))
                e.gender = to_str(gv(r, "Gender", "gender"))
                e.dob = to_date(gv(r, "Date of Birth", "dob"))
                e.race = to_str(gv(r, "Race", "race"))
                e.country = to_str(gv(r, "Country", "country"))
                e.residency = to_str(gv(r, "Residency", "residency"))
                e.pr_date = to_date(gv(r, "PR Date", "pr_date"))
                e.employment_status = to_str(gv(r, "Employment Status", "employment_status"))
                e.employment_pass = to_str(gv(r, "Employment Pass", "employment_pass"))
                e.work_permit_number = to_str(gv(r, "Work Permit Number", "work_permit_number"))
                e.department = to_str(gv(r, "Department", "department"))
                e.position = to_str(gv(r, "Position", "position"))
                e.employment_type = to_str(gv(r, "Employment Type", "employment_type"))
                e.join_date = to_date(gv(r, "Join Date", "join_date"))
                e.exit_date = to_date(gv(r, "Exit Date", "exit_date"))
                e.holiday_group = to_str(gv(r, "Holiday Group", "holiday_group"))
                e.bank = to_str(gv(r, "Bank", "bank"))
                e.bank_account = to_str(gv(r, "Account Number", "bank_account"))

                e.incentives = to_float(gv(r, "Incentives", "incentives"))
                e.allowance = to_float(gv(r, "Allowance", "allowance"))
                e.overtime_rate = to_float(gv(r, "Overtime Rate", "overtime_rate"))
                e.parttime_rate = to_float(gv(r, "Part Time Rate", "parttime_rate"))
                e.levy = to_float(gv(r, "Levy", "levy"))
                b = gv(r, "Basic Salary", "basic_salary")
                if b is not None:
                    try:
                        e.basic_salary = float(b)
                    except Exception:
                        pass

                if is_new:
                    created += 1
                else:
                    updated += 1

            s.commit()

        try:
            self._reload_employees()
        except Exception:
            pass
        else:
            self._notify_employees_changed()
        QMessageBox.information(self, "Import", f"Import complete. Created {created}, Updated {updated}.")

    def _export_employees_template(self):
        if Workbook is None or DataValidation is None or get_column_letter is None:
            QMessageBox.warning(self, "Template", "openpyxl is required. Install 'openpyxl'.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Employee Template", "employee_template.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "Employee Code", "Full Name", "Email", "Contact Number", "Address",
            "ID Type", "ID Number", "Gender", "Date of Birth", "Race", "Country", "Residency", "PR Date",
            "Employment Status", "Employment Pass", "Work Permit Number",
            "Department", "Position", "Employment Type",
            "Join Date", "Exit Date", "Holiday Group",
            "Bank", "Account Number",
            "Incentives", "Allowance", "Overtime Rate", "Part Time Rate", "Levy", "Basic Salary"
        ]
        ws.append(headers)

        with SessionLocal() as s:
            dropdowns = {c: [r.value for r in s.query(DropdownOption)
                             .filter(DropdownOption.account_id == tenant_id(),
                                     DropdownOption.category == c)
                             .order_by(DropdownOption.value).all()]
                         for c in MANAGED_CATEGORIES}
            groups = [g[0] for g in s.query(Holiday.group_code)
                      .filter(Holiday.account_id == tenant_id())
                      .distinct().all()]
        dropdowns["Employment Status"] = ["Active", "Non-Active"]
        dropdowns["Holiday Group"] = groups

        lists = wb.create_sheet("Dropdowns")
        list_col_map = {}
        col_idx = 1
        for cat, vals in sorted(dropdowns.items()):
            list_col_map[cat] = col_idx
            lists.cell(row=1, column=col_idx, value=cat)
            for i, v in enumerate(vals, start=2):
                lists.cell(row=i, column=col_idx, value=_clean_text(v))
            col_idx += 1

        from openpyxl.utils import get_column_letter as _gcl
        header_to_letter = {h: _gcl(i+1) for i, h in enumerate(headers)}

        def add_validation(header_name: str, category: str):
            if category not in list_col_map:
                return
            col = list_col_map[category]
            rng = f"Dropdowns!${_gcl(col)}$2:${_gcl(col)}$2000"
            dv = DataValidation(type="list", formula1=rng, allow_blank=True)
            ws.add_data_validation(dv)
            col_letter = header_to_letter[header_name]
            dv.add(f"{col_letter}2:{col_letter}2000")

        mapping = {
            "ID Type": "ID Type",
            "Gender": "Gender",
            "Race": "Race",
            "Country": "Country",
            "Residency": "Residency",
            "Employment Status": "Employment Status",
            "Employment Pass": "Employment Pass",
            "Department": "Department",
            "Position": "Position",
            "Employment Type": "Employment Type",
            "Bank": "Bank",
            "Holiday Group": "Holiday Group",
        }
        for hdr, cat in mapping.items():
            add_validation(hdr, cat)

        ws.freeze_panes = "A2"
        try:
            wb.save(path)
            QMessageBox.information(self, "Template", "Template saved.")
        except Exception as ex:
            QMessageBox.warning(self, "Template", f"Failed to create template: {ex}")

    def _apply_id_format(self):
        global EMP_CODE_PREFIX, EMP_CODE_ZPAD
        p = self.id_prefix.text().strip() or "EM-"
        z = int(self.zero_pad.value())
        self.id_preview.setText(f"Preview: {p}{1:0{z}d}")

        def extract_num(code: str):
            if not code: return None
            m = re.search(r"(\d+)$", code)
            return int(m.group(1)) if m else None

        with SessionLocal() as s:
            rows = s.query(Employee).filter(Employee.account_id == tenant_id()).order_by(Employee.id.asc()).all()
            used = set()
            next_seq = 1
            for e in rows:
                n = extract_num(e.code or "")
                if n is None or n in used:
                    while next_seq in used:
                        next_seq += 1
                    n = next_seq
                    next_seq += 1
                used.add(n)
                e.code = f"{p}{n:0{z}d}"
            s.commit()

        EMP_CODE_PREFIX, EMP_CODE_ZPAD = p, z
        _save_code_settings(p, z)

        try:
            self._reload_employees()
        except Exception:
            pass
        else:
            self._notify_employees_changed()

        QMessageBox.information(self, "Employee Codes", "Codes updated.")

    # ----- XLSX helpers -----
    def _need_openpyxl(self) -> bool:
        if Workbook is None or load_workbook is None or DataValidation is None:
            QMessageBox.warning(self, "XLSX", "openpyxl is required for XLSX. Install 'openpyxl' and retry.")
            return True
        return False

    def _gather_dropdowns(self) -> dict[str, list[str]]:
        out: dict[str, list[str]] = {}
        with SessionLocal() as s:
            for cat in MANAGED_CATEGORIES:
                vals = [r.value for r in s.query(DropdownOption)
                        .filter(DropdownOption.account_id == tenant_id(),
                                DropdownOption.category == cat)
                        .order_by(DropdownOption.value).all()]
                out[cat] = vals or [""]
        return out

    def _next_employee_code(self, s) -> str:
        prefix, z = EMP_CODE_PREFIX, EMP_CODE_ZPAD
        existing = [r.code for r in s.query(Employee)
                    .filter(Employee.account_id == tenant_id()).all()
                    if r.code and r.code.startswith(prefix)]
        nums = []
        for c in existing:
            m = re.search(r"(\d+)$", c)
            if m:
                try: nums.append(int(m.group(1)))
                except: pass
        nxt = (max(nums) + 1) if nums else 1
        return f"{prefix}{nxt:0{z}d}"


# ---------- Employee Editor ----------
class EmployeeEditor(QDialog):
    def __init__(self, emp_id: int | None = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Employee")
        self._emp_id = emp_id

        lay = QVBoxLayout(self)
        self.tabs = QTabWidget(self)
        lay.addWidget(self.tabs)

        self._build_personal_tab()
        self._build_employment_tab()
        self._build_payment_tab()
        self._build_remuneration_tab()
        self._build_schedule_tab()
        self._build_entitlement_tab()

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel, parent=self)
        bb.accepted.connect(self._save)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)

        self.resize(1200, 760)

        if self._emp_id:
            self._load(self._emp_id)
        else:
            # Add mode: force blank dates and auto-load default leave
            for w in (self.dob, self.pr_date, self.join_date, self.exit_date):
                w.clear()
            self._set_all_fields_blank()
            self._load_defaults_into_entitlements()

    # --- helpers ---
    def _set_all_fields_blank(self):
        from PySide6.QtWidgets import QLineEdit, QComboBox, QPlainTextEdit, QTextEdit, QSpinBox, QDoubleSpinBox

        def _blank_combo(cb: QComboBox):
            if not cb:
                return
            if cb.count() == 0 or cb.itemText(0) not in ("", "-"):
                cb.insertItem(0, "")
            cb.setCurrentIndex(0)

        for name in (
                "full_name", "email", "contact", "address", "id_number",
                "work_permit_number", "bank_account",
                "incentives", "allowance", "ot_rate", "pt_rate", "levy"
        ):
            w = getattr(self, name, None)
            if isinstance(w, QLineEdit):
                w.clear()
            elif isinstance(w, (QPlainTextEdit, QTextEdit)):
                w.clear()
            elif isinstance(w, (QSpinBox, QDoubleSpinBox)):
                try:
                    if hasattr(w, "setSpecialValueText"):
                        w.setSpecialValueText("")
                        w.setValue(w.minimum())
                    else:
                        w.setValue(0)
                except Exception:
                    w.setValue(0)

        for name in (
                "id_type", "gender", "race", "country", "residency",
                "employment_pass", "department",
                "position", "employment_type", "holiday_group", "bank"
        ):
            w = getattr(self, name, None)
            if isinstance(w, QComboBox):
                _blank_combo(w)

        if hasattr(self, "age_lbl"):
            self.age_lbl.setText("-")

    def _opts(self, category: str) -> list[str]:
        with SessionLocal() as s:
            rows = s.query(DropdownOption)\
                    .filter(DropdownOption.account_id == tenant_id(),
                            DropdownOption.category == category)\
                    .order_by(DropdownOption.value).all()
        return [r.value for r in rows]

    def _holiday_groups(self) -> list[str]:
        with SessionLocal() as s:
            rows = s.query(Holiday.group_code)\
                    .filter(Holiday.account_id == tenant_id())\
                    .distinct().all()
        return [r[0] for r in rows]

    def _set_combo_value(self, cb: QComboBox, value: str | None):
        """Ensure blank stays blank, and select exact value when present."""
        txt = (value or "").strip()
        if txt == "":
            if cb.count() == 0 or cb.itemText(0) != "":
                cb.insertItem(0, "")
            cb.setCurrentIndex(0)
            return
        idx = cb.findText(txt, Qt.MatchExactly)
        if idx < 0:
            # keep list clean; do not inject new values here
            # choose blank if available, else 0
            if cb.count() > 0 and cb.itemText(0) == "":
                cb.setCurrentIndex(0)
            else:
                cb.setCurrentIndex(0)
        else:
            cb.setCurrentIndex(idx)

    # --- Personal ---
    def _build_personal_tab(self):
        w = QWidget(); f = QFormLayout(w)
        self.full_name = QLineEdit()
        self.email = QLineEdit()
        self.contact = QLineEdit()
        self.address = QLineEdit()
        self.id_type = QComboBox(); self.id_type.addItem(""); self.id_type.addItems(self._opts("ID Type"))
        self.id_number = QLineEdit()
        self.gender = QComboBox(); self.gender.addItem(""); self.gender.addItems(self._opts("Gender") or ["Male", "Female"])

        # blankable dates with dd/MM/yyyy display and typing
        self.dob = BlankableDateEdit(display_fmt="dd/MM/yyyy")
        self.dob.dateChanged.connect(self._update_age)

        self.age_lbl = QLabel("-")
        self.race = QComboBox(); self.race.addItem(""); self.race.addItems(self._opts("Race"))
        self.country = QComboBox(); self.country.addItem(""); self.country.addItems(self._opts("Country"))
        self.residency = QComboBox(); self.residency.addItem(""); self.residency.addItems(self._opts("Residency") or ["Citizen", "Permanent Resident", "Work Pass"])
        self.residency.currentTextChanged.connect(self._toggle_pr_date)

        self.pr_date = BlankableDateEdit(display_fmt="dd/MM/yyyy")
        self.pr_date.setEnabled(False)

        f.addRow("Full Name", self.full_name)
        f.addRow("Email", self.email)
        f.addRow("Contact Number", self.contact)
        f.addRow("Address", self.address)
        f.addRow("ID Type", self.id_type)
        f.addRow("ID Number", self.id_number)
        f.addRow("Gender", self.gender)
        f.addRow("Date of Birth", self.dob)
        f.addRow("Age", self.age_lbl)
        f.addRow("Race", self.race)
        f.addRow("Country", self.country)
        f.addRow("Residency", self.residency)
        f.addRow("PR Date", self.pr_date)

        self.tabs.addTab(w, "Personal")

    def _update_age(self):
        try:
            d = self.dob.date_or_none()
            if not d:
                self.age_lbl.setText("-")
                return
            dob = d.toPython()
            yrs = int((date.today() - dob).days // 365.25)
            self.age_lbl.setText(str(yrs))
        except Exception:
            self.age_lbl.setText("-")

    def _toggle_pr_date(self, txt: str):
        self.pr_date.setEnabled(txt == "Permanent Resident")

    # sync status rule from exit date
    def _sync_status_from_exit(self):
        qd = self.exit_date
        today = QDate.currentDate()
        is_blank = (qd.date_or_none() is None) if isinstance(qd, BlankableDateEdit) else False
        if is_blank:
            self.employment_status.setCurrentText("Active")
        else:
            if qd.date() > today:
                self.employment_status.setCurrentText("Active")
            elif qd.date() < today:
                self.employment_status.setCurrentText("Non-Active")
            else:
                self.employment_status.setCurrentText("Active")

    # --- Employment ---
    def _build_employment_tab(self):
        w = QWidget()
        f = QFormLayout(w)
        self.employment_status = QComboBox()
        self.employment_status.addItems(["Active", "Non-Active"])
        self.employment_pass = QComboBox()
        self.employment_pass.addItem("")
        self.employment_pass.addItems(self._opts("Employment Pass") or ["None", "S Pass", "Work Permit"])
        self.employment_pass.currentTextChanged.connect(self._toggle_wp)
        self.work_permit_number = QLineEdit()
        self.work_permit_number.setEnabled(False)
        self.department = QComboBox()
        self.department.addItem(""); self.department.addItems(self._opts("Department"))
        self.position = QComboBox()
        self.position.addItem(""); self.position.addItems(self._opts("Position"))
        self.employment_type = QComboBox()
        self.employment_type.addItem(""); self.employment_type.addItems(self._opts("Employment Type") or ["Full-Time", "Part-Time", "Contract"])

        # blankable join and exit dates with dd/MM/yyyy
        self.join_date = BlankableDateEdit(display_fmt="dd/MM/yyyy")
        self.exit_date = BlankableDateEdit(display_fmt="dd/MM/yyyy")
        self.exit_date.dateChanged.connect(lambda _d: self._sync_status_from_exit())

        self.holiday_group = QComboBox()
        self.holiday_group.addItem(""); self.holiday_group.addItems(self._holiday_groups())

        f.addRow("Employment Status", self.employment_status)
        f.addRow("Employment Pass", self.employment_pass)
        f.addRow("Work Permit Number", self.work_permit_number)
        f.addRow("Department", self.department)
        f.addRow("Position", self.position)
        f.addRow("Employment Type", self.employment_type)
        f.addRow("Join Date", self.join_date)
        f.addRow("Exit Date", self.exit_date)
        f.addRow("Holiday Group", self.holiday_group)
        self.tabs.addTab(w, "Employment")

    def _toggle_wp(self, txt: str):
        self.work_permit_number.setEnabled(txt in ("Work Permit", "S Pass"))

    # --- Payment ---
    def _build_payment_tab(self):
        w = QWidget(); f = QFormLayout(w)
        self.bank = QComboBox(); self.bank.addItem(""); self.bank.addItems(self._opts("Bank"))
        self.bank_account = QLineEdit()
        f.addRow("Bank", self.bank)
        f.addRow("Account Number", self.bank_account)
        self.tabs.addTab(w, "Payment")

    # --- Remuneration ---
    def _build_remuneration_tab(self):
        w = QWidget(); v = QVBoxLayout(w)

        grid = QGridLayout()
        self.incentives = QLineEdit("0")
        self.allowance = QLineEdit("0")
        self.ot_rate = QLineEdit("0")
        self.pt_rate = QLineEdit("0")
        self.levy = QLineEdit("0")
        self.basic_salary_lbl = QLabel("0.00")
        grid.addWidget(QLabel("Basic Salary (auto)"), 0, 0); grid.addWidget(self.basic_salary_lbl, 0, 1)
        grid.addWidget(QLabel("Incentives"), 1, 0); grid.addWidget(self.incentives, 1, 1)
        grid.addWidget(QLabel("Allowance"), 2, 0); grid.addWidget(self.allowance, 2, 1)
        grid.addWidget(QLabel("Overtime Rate"), 3, 0); grid.addWidget(self.ot_rate, 3, 1)
        grid.addWidget(QLabel("Part Time Rate"), 4, 0); grid.addWidget(self.pt_rate, 4, 1)
        grid.addWidget(QLabel("Levy"), 5, 0); grid.addWidget(self.levy, 5, 1)
        v.addLayout(grid)

        self.salary_tbl = QTableWidget(0, 3)
        self.salary_tbl.setHorizontalHeaderLabels(["Amount", "Start Date", "End Date"])
        self.salary_tbl.horizontalHeader().setStretchLastSection(True)

        btn_row = QHBoxLayout()
        add = QPushButton("Add Row"); add.clicked.connect(lambda: self._row_add(self.salary_tbl, ["0.00", date.today().strftime("%Y-%m-%d"), ""]))
        rm = QPushButton("Delete Row"); rm.clicked.connect(lambda: self._row_del(self.salary_tbl))
        btn_row.addWidget(add); btn_row.addWidget(rm); btn_row.addStretch(1)

        v.addWidget(QLabel("Salary History"))
        v.addLayout(btn_row)
        v.addWidget(self.salary_tbl, 1)

        self.tabs.addTab(w, "Remuneration")

    # --- Work schedule ---
    def _build_schedule_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        self.ws_tbl = QTableWidget(7, 3)
        self.ws_tbl.setHorizontalHeaderLabels(["Day", "Working", "Day Type"])
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i, d in enumerate(days):
            self.ws_tbl.setItem(i, 0, QTableWidgetItem(d))
            chk = QCheckBox(); chk.setChecked(True); self.ws_tbl.setCellWidget(i, 1, chk)
            typ = QComboBox(); typ.addItems(["Full", "Half"]); self.ws_tbl.setCellWidget(i, 2, typ)
        self.ws_tbl.horizontalHeader().setStretchLastSection(True)
        v.addWidget(self.ws_tbl)
        self.tabs.addTab(w, "Work Schedule")

    # --- Leave entitlement ---
    def _build_entitlement_tab(self):
        w = QWidget(); v = QVBoxLayout(w)

        top = QHBoxLayout()
        self.ent_leave_types = self._load_leave_types()
        self.ent_tbl = QTableWidget(50, max(1, len(self.ent_leave_types)))
        headers = [f"Year {i}" for i in range(1, 51)]
        self.ent_tbl.setVerticalHeaderLabels(headers)
        self.ent_tbl.setHorizontalHeaderLabels(self.ent_leave_types or ["Leave"])
        for r in range(50):
            for c in range(max(1, len(self.ent_leave_types))):
                self.ent_tbl.setItem(r, c, QTableWidgetItem("0"))

        load_def = QPushButton("Load Default Leave Values"); load_def.clicked.connect(self._load_defaults_into_entitlements)
        top.addWidget(load_def); top.addStretch(1)
        v.addLayout(top)
        v.addWidget(self.ent_tbl, 1)
        self.tabs.addTab(w, "Leave Entitlement")

    def _load(self, emp_id: int):
        with SessionLocal() as s:
            e = s.get(Employee, emp_id)
            if not e:
                return
            # Personal
            self.full_name.setText(e.full_name or "")
            self.email.setText(e.email or "")
            self.contact.setText(e.contact_number or "")
            self.address.setText(e.address or "")
            self._set_combo_value(self.id_type, e.id_type)
            self.id_number.setText(e.id_number or "")
            self._set_combo_value(self.gender, e.gender)

            if e.dob:
                self.dob.set_real_date(QDate(e.dob.year, e.dob.month, e.dob.day))
                self._update_age()
            else:
                self.dob.clear()
                self._update_age()
            self._set_combo_value(self.race, e.race)
            self._set_combo_value(self.country, e.country)
            self._set_combo_value(self.residency, e.residency)
            self._toggle_pr_date(self.residency.currentText())
            if e.pr_date:
                self.pr_date.set_real_date(QDate(e.pr_date.year, e.pr_date.month, e.pr_date.day))
            else:
                self.pr_date.clear()

            # Employment
            # keep default behaviour for status via exit-date rule
            self.employment_status.setCurrentText(e.employment_status or "Active")
            self._set_combo_value(self.employment_pass, e.employment_pass)
            self._toggle_wp(self.employment_pass.currentText())
            self.work_permit_number.setText(e.work_permit_number or "")
            self._set_combo_value(self.department, getattr(e, "department", ""))
            self._set_combo_value(self.position, e.position)
            self._set_combo_value(self.employment_type, e.employment_type)

            if e.join_date:
                self.join_date.set_real_date(QDate(e.join_date.year, e.join_date.month, e.join_date.day))
            else:
                self.join_date.clear()

            if e.exit_date:
                self.exit_date.set_real_date(QDate(e.exit_date.year, e.exit_date.month, e.exit_date.day))
            else:
                self.exit_date.clear()
            self._sync_status_from_exit()

            self._set_combo_value(self.holiday_group, e.holiday_group)

            # Payment
            self._set_combo_value(self.bank, e.bank)
            self.bank_account.setText(e.bank_account or "")

            # Remuneration
            self.incentives.setText(str(e.incentives or 0))
            self.allowance.setText(str(e.allowance or 0))
            self.ot_rate.setText(str(e.overtime_rate or 0))
            self.pt_rate.setText(str(e.parttime_rate or 0))
            self.levy.setText(str(e.levy or 0))
            self.basic_salary_lbl.setText(f"{(e.basic_salary or 0.0):.2f}")

            # Salary history
            self.salary_tbl.setRowCount(0)
            sh = (
                s.query(SalaryHistory)
                .filter(SalaryHistory.employee_id == emp_id)
                .order_by(SalaryHistory.start_date.desc())
                .all()
            )
            for row in sh:
                r = self.salary_tbl.rowCount()
                self.salary_tbl.insertRow(r)
                self.salary_tbl.setItem(r, 0, QTableWidgetItem(f"{(row.amount or 0.0):.2f}"))
                self.salary_tbl.setItem(r, 1,
                                        QTableWidgetItem(row.start_date.strftime("%Y-%m-%d") if row.start_date else ""))
                self.salary_tbl.setItem(r, 2,
                                        QTableWidgetItem(row.end_date.strftime("%Y-%m-%d") if row.end_date else ""))

            # Work schedule
            ws = {d.weekday: d for d in s.query(WorkScheduleDay).filter(WorkScheduleDay.employee_id == emp_id).all()}
            for i in range(7):
                chk: QCheckBox = self.ws_tbl.cellWidget(i, 1)
                cmb: QComboBox = self.ws_tbl.cellWidget(i, 2)
                if i in ws:
                    chk.setChecked(ws[i].working)
                    cmb.setCurrentText(ws[i].day_type or "Full")
                else:
                    chk.setChecked(True)
                    cmb.setCurrentText("Full")

            # Entitlements
            ents = s.query(LeaveEntitlement).filter(LeaveEntitlement.employee_id == emp_id).all()
            types = sorted({row.leave_type for row in ents}) or getattr(self, "ent_leave_types", [])
            self.ent_leave_types = list(types) if types else ["Leave"]
            self.ent_tbl.setColumnCount(len(self.ent_leave_types))
            self.ent_tbl.setHorizontalHeaderLabels(self.ent_leave_types)
            grid = {(r.year_of_service, r.leave_type): r.days for r in ents}
            for r in range(50):
                for c, t in enumerate(self.ent_leave_types):
                    val = grid.get((r + 1, t), 0)
                    self.ent_tbl.setItem(r, c, QTableWidgetItem(str(val)))

    def _load_leave_types(self) -> list[str]:
        # combine types from LeaveDefault AND existing per-employee entitlements (tenant-scoped), de-duped case-insensitively
        out = []
        seen = set()
        with SessionLocal() as s:
            for (t,) in s.query(LeaveDefault.leave_type).filter(LeaveDefault.account_id == tenant_id()).distinct().all():
                k = (t or "").strip().casefold()
                if k and k not in seen:
                    seen.add(k); out.append(t)
            for (t,) in s.query(LeaveEntitlement.leave_type).filter(LeaveEntitlement.account_id == tenant_id()).distinct().all():
                k = (t or "").strip().casefold()
                if k and k not in seen:
                    seen.add(k); out.append(t)
        return out or ["Annual Leave"]

    def _load_defaults_into_entitlements(self):
        with SessionLocal() as s:
            rows = s.query(LeaveDefault).filter(LeaveDefault.account_id == tenant_id()).all()
            raw = {d.leave_type: json.loads(d.table_json or "{}") for d in rows}
        if not raw:
            QMessageBox.information(self, "Leave Defaults", "No defaults defined.")
            return

        years_map: dict[str, dict[str, int]] = {}
        for lt, blob in raw.items():
            if isinstance(blob, dict) and "years" in blob and isinstance(blob["years"], dict):
                years_map[lt] = {k: int(v) for k, v in blob["years"].items()}
            elif isinstance(blob, dict):
                years_map[lt] = {k: int(v) for k, v in blob.items()}
            else:
                years_map[lt] = {}

        self.ent_leave_types = list(years_map.keys())
        self.ent_tbl.setColumnCount(len(self.ent_leave_types))
        self.ent_tbl.setHorizontalHeaderLabels(self.ent_leave_types)

        for r in range(50):
            for c, t in enumerate(self.ent_leave_types):
                val = years_map.get(t, {}).get(str(r + 1), 0)
                self.ent_tbl.setItem(r, c, QTableWidgetItem(str(val)))

    # --- utils ---
    def _row_add(self, tbl: QTableWidget, values: list[str] | None = None):
        r = tbl.rowCount(); tbl.insertRow(r)
        vals = values or ["", "", ""]
        for c, v in enumerate(vals):
            tbl.setItem(r, c, QTableWidgetItem(v))

    def _row_del(self, tbl: QTableWidget):
        rows = sorted({i.row() for i in tbl.selectedIndexes()}, reverse=True)
        for r in rows: tbl.removeRow(r)


# ---- Dropdown Options dialog ----
class DropdownOptionsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Dropdown Options")
        v = QVBoxLayout(self)

        top = QHBoxLayout()
        self.cat = QComboBox(); self.cat.addItems(MANAGED_CATEGORIES)
        self.cat.currentTextChanged.connect(self._reload_values)
        btn_add_val = QPushButton("Add"); btn_add_val.clicked.connect(self._add_value)
        btn_rename_val = QPushButton("Rename"); btn_rename_val.clicked.connect(self._rename_value)
        btn_del_val = QPushButton("Delete"); btn_del_val.clicked.connect(self._delete_values)
        for w in (QLabel("Category"), self.cat, btn_add_val, btn_rename_val, btn_del_val):
            top.addWidget(w)
        v.addLayout(top)

        self.val_list = QListWidget()
        v.addWidget(self.val_list, 1)

        self._reload_values(self.cat.currentText())

        bb = QDialogButtonBox(QDialogButtonBox.Close)
        bb.rejected.connect(self.reject); bb.accepted.connect(self.accept)
        v.addWidget(bb)

    def _reload_values(self, category: str):
        self.val_list.clear()
        _ensure_dropdown_defaults(category)
        with SessionLocal() as s:
            rows = s.query(DropdownOption)\
                    .filter(DropdownOption.account_id == tenant_id(),
                            DropdownOption.category == category)\
                    .order_by(DropdownOption.value).all()
        for r in rows:
            QListWidgetItem(r.value, self.val_list)

    def _add_value(self):
        cat = self.cat.currentText()
        txt, ok = self._prompt("Add value", "Value")
        if not ok or not txt.strip(): return
        txt = _clean_text(txt)
        with SessionLocal() as s:
            exists = s.query(DropdownOption).filter(
                DropdownOption.account_id == tenant_id(),
                DropdownOption.category == cat,
                DropdownOption.value == txt.strip()
            ).first()
            if exists:
                QMessageBox.information(self, "Dropdown", "Value already exists."); return
            s.add(DropdownOption(account_id=tenant_id(), category=cat, value=txt.strip()))
            s.commit()
        self._reload_values(cat)

    def _rename_value(self):
        cat = self.cat.currentText()
        it = self.val_list.currentItem()
        if not it: return
        old = it.text()
        if old in DEFAULT_DROPDOWN_OPTIONS.get(cat, []):
            QMessageBox.information(self, "Dropdown", "Default values cannot be renamed.")
            return
        new, ok = self._prompt("Rename value", "New value", old)
        if not ok or not new.strip() or new.strip() == old: return
        new = _clean_text(new)
        with SessionLocal() as s:
            row = s.query(DropdownOption).filter(
                DropdownOption.account_id == tenant_id(),
                DropdownOption.category == cat,
                DropdownOption.value == old
            ).first()
            if not row: return
            exists = s.query(DropdownOption).filter(
                DropdownOption.account_id == tenant_id(),
                DropdownOption.category == cat,
                DropdownOption.value == new.strip()
            ).first()
            if exists:
                QMessageBox.information(self, "Dropdown", "Target value already exists."); return
            row.value = new.strip()
            s.commit()
        self._reload_values(cat)

    def _delete_values(self):
        cat = self.cat.currentText()
        items = self.val_list.selectedItems()
        if not items: return
        protected = set(DEFAULT_DROPDOWN_OPTIONS.get(cat, []))
        blocked = [i.text() for i in items if i.text() in protected]
        if blocked:
            QMessageBox.information(
                self,
                "Dropdown",
                "Default values cannot be deleted:\n- " + "\n- ".join(blocked),
            )
        vals = [i.text() for i in items if i.text() not in protected]
        if not vals:
            return
        if QMessageBox.question(self, "Dropdown", f"Delete {len(vals)} selected value(s)?") != QMessageBox.Yes:
            return
        with SessionLocal() as s:
            for v in vals:
                q = s.query(DropdownOption).filter(
                    DropdownOption.account_id == tenant_id(),
                    DropdownOption.category == cat,
                    DropdownOption.value == v
                )
                for d in q.all(): s.delete(d)
            s.commit()
        self._reload_values(cat)

    def _prompt(self, title: str, label: str, value: str = ""):
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        lay = QVBoxLayout(dlg)
        inp = QLineEdit(value)
        fl = QFormLayout(); fl.addRow(label, inp); lay.addLayout(fl)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(bb)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject)
        if dlg.exec() == QDialog.Accepted:
            return inp.text(), True
        return "", False

# ---- Leave Defaults dialog ----
class LeaveDefaultsDialog(QDialog):
    """
    Stores per-leave settings inside LeaveDefault:
      - columns: prorated (bool)
      - columns: yearly_reset (bool)  <- derived from carry policy: Reset=True, Bring Forward=False
      - table_json: {
            "years": {"1": 14, "2": 14, ...},
            "_meta": {
                "carry_policy": "reset" | "bring",
                "carry_limit_enabled": true|false,
                "carry_limit": 0.0
            }
        }
    Backward compatible with old table_json that was a plain {"1": days,...}.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Leave Defaults")
        self.resize(760, 520)

        root = QHBoxLayout(self)

        # left: type list
        left = QVBoxLayout()
        self.type_list = QListWidget()
        self.type_list.currentItemChanged.connect(self._on_type_changed)
        left.addWidget(QLabel("Leave types"))
        left.addWidget(self.type_list, 1)

        btns = QHBoxLayout()
        add = QPushButton("Add Type"); add.clicked.connect(self._add_type)
        ren = QPushButton("Rename"); ren.clicked.connect(self._rename_type)
        rm  = QPushButton("Delete"); rm.clicked.connect(self._delete_type)
        for b in (add, ren, rm): b.setMaximumWidth(120)
        btns.addWidget(add); btns.addWidget(ren); btns.addWidget(rm); btns.addStretch(1)
        left.addLayout(btns)

        # right: settings + table
        right = QVBoxLayout()

        top = QHBoxLayout()
        self.lbl_curr = QLabel("—")
        top.addWidget(QLabel("Selected:")); top.addWidget(self.lbl_curr); top.addStretch(1)
        right.addLayout(top)

        # --- settings row ---
        meta1 = QHBoxLayout()
        self.prorated = QComboBox(); self.prorated.addItems(["False", "True"])
        meta1.addWidget(QLabel("Pro-rated?")); meta1.addWidget(self.prorated)
        meta1.addSpacing(24)

        self.carry_policy = QComboBox(); self.carry_policy.addItems(["Reset", "Bring forward"])
        self.carry_policy.currentTextChanged.connect(self._toggle_carry_ui)
        meta1.addWidget(QLabel("Year-end handling")); meta1.addWidget(self.carry_policy)
        meta1.addStretch(1)
        right.addLayout(meta1)

        meta2 = QHBoxLayout()
        self.carry_limit_enable = QCheckBox("Limit Bring Forward")
        self.carry_limit_enable.toggled.connect(self._toggle_carry_ui)
        self.carry_limit = QDoubleSpinBox()
        self.carry_limit.setRange(0.0, 365.0)
        self.carry_limit.setDecimals(1)
        self.carry_limit.setSingleStep(0.5)
        self.carry_limit.setSuffix(" day(s)")
        meta2.addWidget(self.carry_limit_enable)
        meta2.addWidget(self.carry_limit)
        meta2.addStretch(1)
        right.addLayout(meta2)

        self.tbl = QTableWidget(50, 2)
        self.tbl.setHorizontalHeaderLabels(["Year", "Days"])
        for i in range(50):
            self.tbl.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.tbl.setItem(i, 1, QTableWidgetItem("14"))
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        right.addWidget(self.tbl, 1)

        save_row = QHBoxLayout()
        save = QPushButton("Save"); save.clicked.connect(self._save_current)
        save.setMaximumWidth(120)
        save_row.addWidget(save); save_row.addStretch(1)
        right.addLayout(save_row)

        root.addLayout(left, 1)
        root.addLayout(right, 2)

        self._load_types()
        self._toggle_carry_ui()

    def _toggle_carry_ui(self):
        bring = (self.carry_policy.currentText() == "Bring forward")
        self.carry_limit_enable.setEnabled(bring)
        self.carry_limit.setEnabled(bring and self.carry_limit_enable.isChecked())

    def _load_types(self):
        self.type_list.clear()
        with SessionLocal() as s:
            rows = s.query(LeaveDefault.leave_type)\
                    .filter(LeaveDefault.account_id == tenant_id())\
                    .distinct().order_by(LeaveDefault.leave_type).all()
        for (t,) in rows:
            self.type_list.addItem(t)
        if self.type_list.count() == 0:
            self.type_list.addItem("Annual Leave")
            self._ensure_row("Annual Leave")
        self.type_list.setCurrentRow(0)

    def _ensure_row(self, leave_type: str):
        with SessionLocal() as s:
            row = s.query(LeaveDefault)\
                   .filter(LeaveDefault.account_id == tenant_id(),
                           LeaveDefault.leave_type == leave_type)\
                   .first()
            if not row:
                years = {str(i + 1): 14 for i in range(50)}
                meta = {"carry_policy": "reset", "carry_limit_enabled": False, "carry_limit": 0.0}
                s.add(LeaveDefault(
                    account_id=tenant_id(), leave_type=leave_type,
                    prorated=False, yearly_reset=True,
                    table_json=json.dumps({"years": years, "_meta": meta})
                ))
                s.commit()

    def _on_type_changed(self, curr, _prev):
        if not curr:
            return
        typ = curr.text()
        self.lbl_curr.setText(typ)
        with SessionLocal() as s:
            row = s.query(LeaveDefault)\
                   .filter(LeaveDefault.account_id == tenant_id(),
                           LeaveDefault.leave_type == typ).first()
        if not row:
            self._ensure_row(typ)
            with SessionLocal() as s:
                row = s.query(LeaveDefault)\
                       .filter(LeaveDefault.account_id == tenant_id(),
                               LeaveDefault.leave_type == typ).first()

        self.prorated.setCurrentText("True" if row.prorated else "False")

        try:
            blob = json.loads(row.table_json or "{}")
        except Exception:
            blob = {}

        if "years" in blob and isinstance(blob["years"], dict):
            years = blob["years"]
            meta = blob.get("_meta", {})
        else:
            years = blob if isinstance(blob, dict) else {}
            meta = {}

        carry_policy = meta.get("carry_policy", "reset")
        self.carry_policy.setCurrentText("Bring forward" if carry_policy == "bring" else "Reset")
        self.carry_limit_enable.setChecked(bool(meta.get("carry_limit_enabled", False)))
        try:
            self.carry_limit.setValue(float(meta.get("carry_limit", 0.0)))
        except Exception:
            self.carry_limit.setValue(0.0)

        row.yearly_reset = (carry_policy != "bring")
        with SessionLocal() as s:
            s.merge(row); s.commit()

        for i in range(50):
            self.tbl.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.tbl.setItem(i, 1, QTableWidgetItem(str(years.get(str(i + 1), 14))))

        self._toggle_carry_ui()

    def _save_current(self):
        itm = self.type_list.currentItem()
        if not itm:
            return
        typ = itm.text()

        years = {str(i + 1): self._int_or_zero(self.tbl.item(i, 1)) for i in range(50)}
        carry_policy = "bring" if self.carry_policy.currentText() == "Bring forward" else "reset"
        meta = {
            "carry_policy": carry_policy,
            "carry_limit_enabled": self.carry_limit_enable.isChecked(),
            "carry_limit": float(self.carry_limit.value()),
        }

        with SessionLocal() as s:
            row = s.query(LeaveDefault)\
                   .filter(LeaveDefault.account_id == tenant_id(),
                           LeaveDefault.leave_type == typ).first()
            if not row:
                row = LeaveDefault(account_id=tenant_id(), leave_type=typ)
                s.add(row)

            row.prorated = (self.prorated.currentText() == "True")
            row.yearly_reset = (carry_policy != "bring")
            row.table_json = json.dumps({"years": years, "_meta": meta})
            s.commit()

        QMessageBox.information(self, "Leave Defaults", f"Saved '{typ}'")

    def _add_type(self):
        name, ok = QInputDialog.getText(self, "Add Leave Type", "Name:")
        if not ok or not name.strip():
            return
        name = _clean_text(name)
        self._ensure_row(name)
        self._load_types()
        items = self.type_list.findItems(name, Qt.MatchExactly)
        if items:
            self.type_list.setCurrentItem(items[0])

    def _rename_type(self):
        itm = self.type_list.currentItem()
        if not itm:
            return
        old = itm.text()
        new, ok = QInputDialog.getText(self, "Rename Leave Type", "New name:", text=old)
        if not ok or not new.strip() or new == old:
            return
        new = _clean_text(new)
        with SessionLocal() as s:
            row = s.query(LeaveDefault)\
                   .filter(LeaveDefault.account_id == tenant_id(),
                           LeaveDefault.leave_type == old).first()
            if row:
                row.leave_type = new
                s.commit()
        self._load_types()
        items = self.type_list.findItems(new, Qt.MatchExactly)
        if items:
            self.type_list.setCurrentItem(items[0])

    def _delete_type(self):
        itm = self.type_list.currentItem()
        if not itm:
            return
        typ = itm.text()
        if QMessageBox.question(self, "Delete", f"Delete leave type '{typ}'?") != QMessageBox.Yes:
            return
        with SessionLocal() as s:
            rows = s.query(LeaveDefault)\
                    .filter(LeaveDefault.account_id == tenant_id(),
                            LeaveDefault.leave_type == typ).all()
            for r in rows:
                s.delete(r)
            s.commit()
        self._load_types()

    @staticmethod
    def _int_or_zero(item: QTableWidgetItem | None) -> int:
        try:
            return int((item.text() if item else "0").strip() or "0")
        except Exception:
            return 0


# ---------------- EmployeeEditor save/load (at end to keep file compact) ----------------
def _extract_trailing_int(txt: str | None):
    if not txt:
        return None
    m = re.search(r"(\d+)$", txt)
    return int(m.group(1)) if m else None


def _employeeeditor_save(self: EmployeeEditor):
    def f2(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    def dget(qd: QDateEdit) -> date | None:
        try:
            if hasattr(qd, "date_or_none"):
                d = qd.date_or_none()
                return d.toPython() if d else None
            if qd.specialValueText() and qd.date() == qd.minimumDate():
                return None
            return qd.date().toPython()
        except Exception:
            return None

    name = _clean_text(self.full_name.text())
    if not name:
        QMessageBox.warning(self, "Missing", "Full Name is required.")
        return

    with SessionLocal() as s:
        if self._emp_id:
            e = s.get(Employee, self._emp_id)
            if not e:
                QMessageBox.critical(self, "Error", "Employee record not found.")
                return
        else:
            prefix, z = EMP_CODE_PREFIX, EMP_CODE_ZPAD
            existing = [
                c for (c,) in s.query(Employee.code)
                .filter(Employee.account_id == tenant_id(), Employee.code.isnot(None), Employee.code.like(f"{prefix}%"))
                .all()
            ]
            def _num_tail(c):
                if not c: return None
                m = re.search(r"(\d+)$", c)
                return int(m.group(1)) if m else None
            used = {n for n in (_num_tail(c) for c in existing) if n is not None}
            nxt = (max(used) + 1) if used else 1
            code = f"{prefix}{nxt:0{z}d}"

            e = Employee(account_id=tenant_id(), code=code, full_name=name)
            s.add(e)
            s.flush()
            self._emp_id = e.id

        # personal
        e.full_name = name
        e.email = _clean_text(self.email.text())
        e.contact_number = _clean_text(self.contact.text())
        e.address = _clean_text(self.address.text())
        e.id_type = self.id_type.currentText() or ""
        e.id_number = _clean_text(self.id_number.text())
        e.gender = self.gender.currentText() or ""
        e.dob = dget(self.dob)
        e.race = self.race.currentText() or ""
        e.country = self.country.currentText() or ""
        e.residency = self.residency.currentText() or ""
        e.pr_date = dget(self.pr_date) if self.pr_date.isEnabled() else None

        # employment
        e.employment_status = self.employment_status.currentText() or ""
        e.employment_pass = self.employment_pass.currentText() or ""
        e.work_permit_number = _clean_text(self.work_permit_number.text()) if self.work_permit_number.isEnabled() else ""
        e.department = self.department.currentText() or ""
        e.position = self.position.currentText() or ""
        e.employment_type = self.employment_type.currentText() or ""
        e.join_date = dget(self.join_date)
        e.exit_date = dget(self.exit_date)
        e.holiday_group = self.holiday_group.currentText() or ""

        # payment
        e.bank = self.bank.currentText() or ""
        e.bank_account = _clean_text(self.bank_account.text())

        # remuneration snapshot
        e.incentives = f2(self.incentives.text())
        e.allowance = f2(self.allowance.text())
        e.overtime_rate = f2(self.ot_rate.text())
        e.parttime_rate = f2(self.pt_rate.text())
        e.levy = f2(self.levy.text())

        # salary history -> compute basic from latest start_date
        s.query(SalaryHistory).filter(SalaryHistory.employee_id == e.id).delete()
        latest_amt, latest_start = 0.0, date(1900, 1, 1)
        for r in range(self.salary_tbl.rowCount()):
            try:
                amt = f2(self.salary_tbl.item(r, 0).text())
                sd_txt = _clean_text(self.salary_tbl.item(r, 1).text() or "")
                ed_txt = _clean_text(self.salary_tbl.item(r, 2).text() or "")
                sd = datetime.strptime(sd_txt, "%Y-%m-%d").date() if sd_txt else None
                ed = datetime.strptime(ed_txt, "%Y-%m-%d").date() if ed_txt else None
            except Exception:
                continue
            row = SalaryHistory(account_id=tenant_id(), employee_id=e.id, amount=amt, start_date=sd, end_date=ed)
            s.add(row)
            if sd and sd >= latest_start:
                latest_start = sd
                latest_amt = amt
        e.basic_salary = latest_amt

        # work schedule
        s.query(WorkScheduleDay).filter(WorkScheduleDay.employee_id == e.id).delete()
        for i in range(7):
            chk: QCheckBox = self.ws_tbl.cellWidget(i, 1)
            cmb: QComboBox = self.ws_tbl.cellWidget(i, 2)
            s.add(WorkScheduleDay(
                account_id=tenant_id(), employee_id=e.id,
                weekday=i, working=chk.isChecked(), day_type=cmb.currentText()
            ))

        # entitlements
        s.query(LeaveEntitlement).filter(LeaveEntitlement.employee_id == e.id).delete()
        for r in range(50):
            for c in range(self.ent_tbl.columnCount()):
                hdr = self.ent_tbl.horizontalHeaderItem(c)
                t = hdr.text() if hdr else "Leave"
                cell = self.ent_tbl.item(r, c)
                try:
                    days = float(cell.text()) if cell and cell.text() else 0.0
                except Exception:
                    days = 0.0
                s.add(LeaveEntitlement(
                    account_id=tenant_id(), employee_id=e.id,
                    year_of_service=r + 1, leave_type=t, days=days
                ))

        try:
            s.commit()
        except Exception as ex:
            s.rollback()
            QMessageBox.critical(self, "Save failed", f"{ex}")
            return

    self.accept()


# bind the save override
EmployeeEditor._save = _employeeeditor_save
